import telebot
import openpyxl
from telebot import types

BOT_TOKEN = '6562890122:AAGRPmeOMeBhs4PJU2Lg7vefahxXoNSppKo'
bot = telebot.TeleBot(BOT_TOKEN)
user_state = {}

wb = openpyxl.Workbook()
sheet = wb.active

# Добавление заголовков
headers = ["ФИО", "Статус", "Дата", "Сумма", "Номер Карты"]
sheet.append(headers)
@bot.message_handler(commands=['start'])
def start(message):
    if message.chat.id in user_state:
        del user_state[message.chat.id]
        bot.send_message(message.chat.id, "Состояние сброшено.")
    markup = types.InlineKeyboardMarkup(row_width=1)
    button_add_customer = types.InlineKeyboardButton("Добавить покупателя", callback_data="add_consumer_")
    button_find_customer = types.InlineKeyboardButton("Найти покупателя", callback_data="find_consumer_")
    show_one_page = types.InlineKeyboardButton("Скачать таблицу", callback_data="download_file")
    button_download_file = types.InlineKeyboardButton("Посмотреть страницу с заказами", callback_data="show_pages_")
    # button_edit_purchase = types.InlineKeyboardButton(text="Изменить данные покупки", callback_data="edit_purchase")
    add_page_button = types.InlineKeyboardButton("Добавить сайт", callback_data='add_page')
    markup.add(show_one_page,button_add_customer, button_find_customer, add_page_button,button_download_file)
    bot.reply_to(message, "Бот заказов! Выберите действие:", reply_markup=markup)
    markup = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    start_button = telebot.types.KeyboardButton(text='/start')
    markup.add(start_button)


@bot.message_handler(commands=['cancel'])
def cancel_handler(message):
    user_id = message.from_user.id
    print(user_state)
    if user_id in user_state:
        del user_state[user_id]
        bot.send_message(user_id, "Действие успешно отменено.")
        start(message)  # Вызываем обработчик команды /start после успешной отмены
        bot.send_message(message.chat.id, "Действие отменено. Выберите действие:", reply_markup=start())

    else:
        bot.reply_to(message, "Нет активных действий для отмены.")
@bot.message_handler(commands=['status'])
def status_handler(message):
    user_id = message.from_user.id
    user_data = user_state.get(user_id, {})
    current_state = user_data.get("state", "Нет состояния")
    bot.reply_to(message, f"Текущее состояние: {current_state}")
@bot.message_handler(func=lambda message: True) # для чего то важного, доделать, работает не корректно
def check_state(message):
    user_id = message.from_user.id
    user_data = user_state.get(user_id, {})

    if user_data.get("state") == "add_customer":
        bot.send_message(user_id, "Вы находитесь в режиме добавления покупателя. Введите данные или отправьте /cancel для отмены.")
    elif user_data.get("state") == "find_customer":
        bot.send_message(user_id, "Вы находитесь в режиме поиска покупателя. Введите ФИО или отправьте /cancel для отмены.")
    else:
        bot.send_message(user_id, "Выберите действие:", reply_markup=start(message))



@bot.callback_query_handler(func=lambda call: call.data.startswith('download_file'))
def download_file(call):
    try:
        user_data = user_state.get(call.message.chat.id, {})
        if user_data.get("state") in ["add_customer", "find_customer"]:
            bot.send_message(call.message.chat.id, "Сначала завершите текущее действие.")
            return

        document_path = 'example.xlsx'  # Путь к вашему файлу
        with open(document_path, 'rb') as document:
            bot.send_document(call.message.chat.id, document)
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Произошла ошибка: {e}")
@bot.callback_query_handler(func=lambda call: call.data.startswith('show_page_'))
def display_page_inline(call):
    if call.from_user.id in user_state:
        bot.send_message(call.message.chat.id, "Вы уже выполняете другую операцию. Дождитесь ее завершения.")
        return
    sheet_name = call.data.replace('display_page_', '')
    user_state[call.message.chat.id] = {"state": "display_page", "sheet_name": sheet_name}
    bot.send_message(call.message.chat.id, "Отображение страницы по столбцам...")
    display_page(call.message)
@bot.callback_query_handler(func=lambda call: call.data.startswith("add_customer_"))
def add_customer_on_page(call):
    if call.from_user.id in user_state:
        bot.send_message(call.message.chat.id, "Вы уже выполняете другую операцию. Дождитесь ее завершения.")
        return

    sheet_name = call.data.replace('add_customer_', '')
    msg = bot.send_message(call.message.chat.id, '''
    Введите запись в формате: 'ФИО Статус Дата Сумма Номер_Карты'
    Пример: Журавлева Оплачено 12.08 4500  3434''')
    user_state[call.from_user.id] = {"state": "add_customer", "sheet_name": sheet_name}
    bot.register_next_step_handler(msg, process_add_customer)
@bot.callback_query_handler(func=lambda call: call.data.startswith('find_customer_'))
def find_customer_on_page(call):
    if call.from_user.id in user_state:
        bot.send_message(call.message.chat.id, "Вы уже выполняете другую операцию. Дождитесь ее завершения.")
        return

    sheet_name = call.data.replace('find_customer_', '')
    msg = bot.send_message(call.message.chat.id, "Введите ФИО покупателя для поиска:")
    user_state[call.from_user.id] = {"state": "find_customer", "sheet_name": sheet_name}
    bot.register_next_step_handler(msg, process_find_customer)




@bot.callback_query_handler(func=lambda call: call.data.startswith('add_consumer_'))
def add_customer_request(call):
    if call.from_user.id in user_state:
        bot.send_message(call.message.chat.id, "Вы уже выполняете другую операцию. Дождитесь ее завершения.")
        return

    try:
        workbook = openpyxl.load_workbook("example.xlsx")
        sheet_names = workbook.sheetnames
        markup = telebot.types.InlineKeyboardMarkup()
        for sheet_name in sheet_names:
            button = telebot.types.InlineKeyboardButton(sheet_name, callback_data=f'add_customer_on_page_{sheet_name}')
            markup.add(button)
        bot.send_message(call.message.chat.id, "Выберите страницу для добавления:", reply_markup=markup)
        workbook.close()
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Произошла ошибка: {e}")
@bot.callback_query_handler(func=lambda call: call.data.startswith('find_consumer_'))
def find_customer_request(call):
    if call.from_user.id in user_state:
        bot.send_message(call.message.chat.id, "Вы уже выполняете другую операцию. Дождитесь ее завершения.")
        return

    try:
        workbook = openpyxl.load_workbook("example.xlsx")
        sheet_names = workbook.sheetnames
        markup = telebot.types.InlineKeyboardMarkup()
        for sheet_name in sheet_names:
            button = telebot.types.InlineKeyboardButton(sheet_name, callback_data=f'find_customer_on_page_{sheet_name}')
            markup.add(button)
        bot.send_message(call.message.chat.id, "Выберите страницу для поиска:", reply_markup=markup)
        workbook.close()
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Произошла ошибка: {e}")
@bot.callback_query_handler(func=lambda call: call.data.startswith('show_pages_'))
def display_pages_columns(call):
    if call.from_user.id in user_state:
        bot.send_message(call.message.chat.id, "Вы уже выполняете другую операцию. Дождитесь ее завершения.")
        return

    try:
        workbook = openpyxl.load_workbook("example.xlsx")
        sheet_names = workbook.sheetnames
        markup = telebot.types.InlineKeyboardMarkup()
        for sheet_name in sheet_names:
            button = telebot.types.InlineKeyboardButton(sheet_name, callback_data=f'show_page_{sheet_name}')
            markup.add(button)
        bot.send_message(call.message.chat.id, "Выберите страницу для просмотра Содержимого:", reply_markup=markup)
        workbook.close()
    except Exception as e:
        bot.send_message(call.message.chat.id, f"Произошла ошибка: {e}")



# ----------------------------------------------------------- добавить страницу

@bot.callback_query_handler(func=lambda call: call.data.startswith('add_page'))
def add_page_inline(call):
    msg = bot.send_message(call.message.chat.id, "Введите имя новой страницы:")
    user_state[call.message.chat.id] = {"state": "add_page"}
    bot.register_next_step_handler(msg, add_page_to_excel)
def add_page_to_excel(message):
    user_data = user_state.get(message.chat.id, {})
    data = message.text.split()
    if user_data.get("state") == "add_page":
        new_sheet_name = message.text
        try:
            if data[0] == '/cancel':
                del user_state[message.from_user.id]
                bot.reply_to(message, "Действие успешно отменено.")
                return
            workbook = openpyxl.load_workbook("example.xlsx")
            new_sheet = workbook.create_sheet(new_sheet_name)
            column_headers = ["ФИО", "Оплачено", "Дата", "Сумма", "НомерКарты"]
            new_sheet.append(column_headers)

            workbook.save("example.xlsx")

            workbook.close()
            bot.send_message(message.chat.id, f"Страница '{new_sheet_name}' успешно добавлена в Excel файл")
        except Exception as e:
            bot.send_message(message.chat.id, f"Произошла ошибка: {e}")
            bot.send_message(message.chat.id, "/start")
        finally:
            del user_state[message.chat.id]
            return
    else:
        bot.send_message(message.chat.id, "Ошибка: неправильное состояние пользователя.")
        bot.send_message(message.chat.id, "/start")
        return
# ----------------------------------------------------------- добавить страницу



# добавить юзера
def process_add_customer(message):
    try:
        user_data = user_state.get(message.from_user.id, {})
        if user_data.get("state") == "add_customer":
            sheet_name = user_data.get("sheet_name").replace("on_page_", "")
            wb = openpyxl.load_workbook('example.xlsx')
            sheet = wb[sheet_name]
            data = message.text.split()

            if len(data) == 1 and data[0] == '/cancel':
                del user_state[message.from_user.id]
                bot.reply_to(message, "Действие успешно отменено.")
                return

            if len(data) == 5:
                sheet.append(data)
                wb.save('example.xlsx')
                bot.reply_to(message, "Данные покупателя успешно добавлены в файл.")
                del user_state[message.from_user.id]
            else:
                bot.reply_to(message, "Ошибка: Не удалось разобрать сообщение. Введите данные в правильном формате или отправьте /cancel для отмены.")
                # Рекурсивно вызываем функцию обработки для ожидания нового ввода
                msg = bot.send_message(message.chat.id, '''
                'ФИО Статус Дата Сумма Номер_Карты'
                Пример: Журавлева Оплачено 12 августа 4500  3434
                Отправьте /cancel для отмены.''')
                bot.register_next_step_handler(msg, process_add_customer)
        else:
            bot.reply_to(message, "Ошибка: неправильное состояние пользователя.")
            bot.send_message(message.chat.id, "/start")
            return
    except Exception as e:
        bot.reply_to(message, "Произошла ошибка: " + str(e))
# найти юзера
def process_find_customer(message):
    try:
        data = message.text.split()
        user_data = user_state.get(message.from_user.id, {})
        if user_data.get("state") == "find_customer":
            sheet_name = user_data.get("sheet_name").replace("on_page_", "")
            wb = openpyxl.load_workbook('example.xlsx')
            sheet = wb[sheet_name]
            fio_to_find = message.text.strip()

            if len(data) == 1 and data[0] == '/cancel':
                del user_state[message.from_user.id]
                bot.reply_to(message, "Действие успешно отменено.")
                return

            found_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if fio_to_find.lower() in row[0].lower():
                    found_data.append("\t".join(str(cell) for cell in row))

            if found_data:
                if len(found_data) == 1:
                    result = "\n".join(found_data)
                    bot.reply_to(message, "Найденные данные:\n" + result)
                else:
                    response = "Найденные данные:\n\n"
                    for data in found_data:
                        response += f"{data}\n\n"
                        if len(response) >= 4000:
                            bot.send_message(message.chat.id, response)
                            response = ""
                    if response:
                        bot.send_message(message.chat.id, response)
            else:
                bot.reply_to(message, "Покупатель не найден.")
            del user_state[message.from_user.id]
        else:
            bot.reply_to(message, "Ошибка: неправильное состояние пользователя.")
            bot.send_message(message.chat.id, "/start")
            return
    except Exception as e:
        bot.reply_to(message, "Произошла ошибка: " + str(e))




def display_page(message):
    user_data = user_state.get(message.chat.id, {})
    if user_data.get("state") == "display_page":
        sheet_name = user_data.get("sheet_name").replace("show_page_", "")
        data = message.text.split()
        try:
            if data[0] == '/cancel':
                del user_state[message.from_user.id]
                bot.reply_to(message, "Действие успешно отменено.")
                return
            workbook = openpyxl.load_workbook('example.xlsx')
            sheet = workbook[sheet_name]

            column_headers = sheet[1]
            column_names = [cell.value for cell in column_headers]

            rows_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows_data.append(row)

            if rows_data:
                response = "Содержимое страницы по строкам:\n\n"
                for row_data in rows_data:
                    row_values = "\n".join(f"{column}: {value}" for column, value in zip(column_names, row_data))
                    response += f"{row_values}\n\n"
                bot.send_message(message.chat.id, response)

            else:
                bot.send_message(message.chat.id, "Страница пуста.")
                return

            del user_state[message.chat.id]  # Удаление состояния пользователя
        except Exception as e:
            bot.send_message(message.chat.id, f"Произошла ошибка: {e}")
            bot.send_message(message.chat.id, "/start")

        finally:
            del user_state[message.chat.id]
            return
    else:
        bot.send_message(message.chat.id, "Ошибка: неправильное состояние пользователя.")
        bot.send_message(message.chat.id, "/start")
        return





# Основной цикл для получения обновлений
def main():
    bot.polling(none_stop=True)


if __name__ == "__main__":
    main()
